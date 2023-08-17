from base.models import Solped, SolpedItem, Product, ObservationsSolped, Category, User, Warehouse, ShippingAddress, Document
from base.serializers import SolpedSerializer, SolpedItemsSerializer, ObservationsSolpedSerializer, ProductSerializer, DocumentSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework import status
import openpyxl
from django.http import HttpResponse
from ..models.solped_models import status_CHOICES
from datetime import datetime
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from django.core.paginator import Paginator
from django.db.models.functions import Cast
from django.db.models import DateTimeField

@api_view(['POST'])
@permission_classes([])
def create_solped(request):
    # Crear solped
    total_price = 0
    solped_serializer = SolpedSerializer(data=request.data, partial=True)
    if solped_serializer.is_valid():
        #solped = solped_serializer.save(creator_user=request.user)
        solped = solped_serializer.save(creator_user=User.objects.get(pk=request.user.pk))
        # Crear items
        items_data = request.data.get('items')
        for item_data in items_data:
            product = Product.objects.get(pk=item_data.get('product_id'))
            SolpedItem.objects.create(
                solped=solped,
                product=product,
                quantity=item_data.get('quantity'),
                unit_price=item_data.get('unit_price'),
                price=item_data.get('quantity') * item_data.get('unit_price')
            )
            total_price += item_data.get('quantity') * item_data.get('unit_price')
        solped_serializer.save(total_price=total_price)
        return Response(solped_serializer.data, status=status.HTTP_201_CREATED)
    else:
        return Response(solped_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([])
def get_solped(request, pk):
    solped = Solped.objects.get(pk=pk)
    if solped is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        solped_serializer = SolpedSerializer(solped, many=False)
        return Response(solped_serializer.data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([])
def get_solpeds(request):
    solpeds = Solped.objects.all()
    if solpeds is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        solped_serializer = SolpedSerializer(solpeds, many=True)
        return Response(solped_serializer.data, status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([])
def update_solped(request, pk):
    solped = Solped.objects.get(pk=pk)
    if solped is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        solped_serializer = SolpedSerializer(instance=solped, data=request.data, partial=True)
        if solped_serializer.is_valid():
            solped_serializer.save()
            return Response(solped_serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(solped_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([])
def delete_solped(request, pk):
    solped = Solped.objects.get(pk=pk)
    if solped is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        solped.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([])
def get_solped_items(request, pk):
    solped = Solped.objects.get(pk=pk)
    if solped is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        solped_items = SolpedItem.objects.filter(solped=solped)
        solped_items_serializer = SolpedItemsSerializer(solped_items, many=True)
        return Response(solped_items_serializer.data, status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([])
def update_solped_items(request, pk):
    solped = Solped.objects.get(pk=pk)
    solped_items = SolpedItem.objects.filter(solped=solped)
    solped_items.delete()
    total_price = 0
    if solped is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        items_data = request.data.get('items')
        for item_data in items_data:
            product = Product.objects.get(pk=item_data.get('product_id'))
            SolpedItem.objects.create(
                solped=solped,
                product=product,
                quantity=item_data.get('quantity'),
                unit_price=item_data.get('unit_price'),
                price=item_data.get('quantity') * item_data.get('unit_price')
            )
            total_price += item_data.get('quantity') * item_data.get('unit_price')
        solped_serializer = SolpedSerializer(instance=solped, data=request.data, partial=True)
        if solped_serializer.is_valid():
            solped_serializer.save(total_price=total_price)
            return Response(solped_serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(solped_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([])
def get_solped_excel(request, pk):
    solped = Solped.objects.get(pk=pk)
    if solped is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        # Crear un nuevo archivo de Excel
        workbook = openpyxl.Workbook()

        # Agregar una hoja al archivo de Excel
        worksheet = workbook.active
        worksheet.title = 'Solped'

        # Agregar los encabezados de columna
        columns = ['Codigo de Material', 'Descripcion del Material', 'Categoria', 'Numero de SOLPED', 'Centro de Costos', 'Usuario Creador de Solped', 'Cantidades',
                   'Unidades de Medida', 'Valor Unitario', 'Valor Total', 'Comprador', 'Estado Actual', 'Fecha de Creación de la SOLPED', 'Fecha de Solución de la SOLPED',
                   'Lugar de Requerimiento', 'Almacen', 'Observaciones']
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        items = SolpedItem.objects.filter(solped=solped)
        observation = ObservationsSolped.objects.filter(solped=solped)
        try:
            warehouse = Warehouse.objects.get(pk=solped.warehouse)
            address = ShippingAddress.objects.get(pk=warehouse.shipping_address)
            warehouse = warehouse.name
            address = address.address
        except:
            warehouse = 'No hay almacen'
            address = 'No hay direccion'

        observations = ''
        num = 1
        for obs in observation:
            observations += f'{num}.'
            observations += obs.observation
            observations += '  '
            num += 1

        # Agregar los datos de la Solped por producto
        if items:
            for item in items:
                category = Category.objects.get(pk=item.product.category_id)
                row_num += 1
                row = [
                    item.product.reference_code if item.product.reference_code else 'No hay codigo de material',
                    item.product.description if item.product.description else 'No hay descripcion',
                    category.name,
                    solped.pk,
                    solped.cost_center if solped.cost_center else 'No hay centro de costo',
                    solped.creator_user.username if solped.creator_user else 'No hay creador',
                    item.quantity if item.quantity else 'No hay cantidad',
                    item.product.measurement_unit if item.product.measurement_unit else 'No hay unidad de medida',
                    item.unit_price if item.unit_price else 'No hay precio',
                    item.price if item.price else 'No hay precio',
                    solped.assigned_negotiator.username if solped.assigned_negotiator else 'No hay comprador',
                    status_CHOICES[solped.status-1][1] if solped.status else 'No hay estado actual',
                    solped.createdAt.strftime("%Y-%m-%d, %H:%M:%S") if solped.createdAt else 'No hay fecha de creacion',
                    solped.authorization_date.strftime("%Y-%m-%d, %H:%M:%S") if solped.authorization_date else 'No hay fecha de solucion',
                    address,
                    warehouse,
                    observations if observations else 'No hay observaciones',
                ]
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value

            # Configurar las dimensiones de la columna
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

            # Guardar el archivo de Excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=solped{solped.pk}.xlsx'
            workbook.save(response)

            return response


@api_view(['PUT'])
@permission_classes([])
def authorization_solped(request, pk):
    solped = Solped.objects.get(pk=pk)
    if solped is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    #if request.user != solped.assigned_negotiator:
    #    return Response({'error': 'No tiene permisos para realizar esta accion'}, status=status.HTTP_401_UNAUTHORIZED)
    solped.authorization_date = datetime.now()
    solped.status = 2
    solped_serializer = SolpedSerializer(solped, data=request.data, partial=True)
    if solped_serializer.is_valid():
        solped_serializer.save()
        return Response(solped_serializer.data, status=status.HTTP_200_OK)
    else:
        return Response(solped_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([])
def create_observation(request):
    user = User.objects.get(pk=request.user.pk)
    solped = Solped.objects.get(pk=request.data.get('solped'))
    if user == solped.assigned_negotiator:
        observation_solped = ObservationsSolped.objects.create(
            owner=user,
            solped=solped,
            observation=request.data.get('observation'),
        )
        observation_solped_serializer = ObservationsSolpedSerializer(observation_solped)
        return Response(observation_solped_serializer.data, status=status.HTTP_201_CREATED)
    else:
        return Response({'error': 'No tiene permisos para realizar esta accion'}, status=status.HTTP_401_UNAUTHORIZED)


from rest_framework.pagination import PageNumberPagination

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_solpeds_by_user(request):

    filter_params = request.query_params
    solpeds = Solped.objects.filter(creator_user=request.user.pk, status__lt=6)

    if solpeds is None:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if 'order' in filter_params:
        order_by = filter_params['order']
        if order_by == 'fecha_creacion':
            solpeds = solpeds.order_by('createdAt')
        elif order_by == 'numero_solped':
            solpeds = solpeds.order_by('_id')
        elif order_by == 'precio_total':
            solpeds = solpeds.order_by('total_price')
        elif order_by == 'fecha_resolucion':
            solpeds = solpeds.order_by('resolution_deadline')
        elif order_by == 'estado':
            solpeds = solpeds.order_by('status')
        else:
            solpeds = solpeds.order_by('createdAt')

    if filter_params['type'] and filter_params['input']:
        type_filter = filter_params['type']
        input_filter = filter_params['input']

        if type_filter == 'numero_solped':
            solpeds = solpeds.filter(_id__contains=input_filter) #works
        elif type_filter == 'codigo_material':
            solpeds = solpeds.filter(solpeditem__product__reference_code__exact=input_filter) #works
        elif type_filter == 'usuario_creador':
            solpeds = solpeds.filter(creator_user=input_filter) # filtrar por nombre
        elif type_filter == 'ubicacion':
            solpeds = solpeds.filter()
        elif type_filter == 'fecha_creacion':
            input_date = datetime.strptime(input_filter, '%Y-%m-%d')
            input_year = input_date.year
            input_month = input_date.month
            input_day = input_date.day
            solpeds = solpeds.filter(createdAt__year=input_year, createdAt__month=input_month, createdAt__day=input_day)
        elif type_filter == 'centro_costos':
            solpeds = solpeds.filter()
        elif type_filter == 'estado':
            solpeds = solpeds.filter(status=int(input_filter))

    # Aplicar paginación
    paginator = PageNumberPagination()
    paginator.page_size =4  # Número de elementos por página
    result_page = paginator.paginate_queryset(solpeds, request)

    solpeds_serializer = SolpedSerializer(result_page, many=True)
    return paginator.get_paginated_response({'solpeds': solpeds_serializer.data, 'pages': paginator.page.paginator.num_pages})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_odc_by_user(request):
    filter_params = request.query_params

    solpeds = Solped.objects.filter(creator_user=request.user.pk, status__gt=4)

    if solpeds is None:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if 'order' in filter_params:
        order_by = filter_params['order']
        if order_by == 'fecha_creacion':
            solpeds = solpeds.order_by('createdAt')
        elif order_by == 'numero_solped':
            solpeds = solpeds.order_by('_id')
        elif order_by == 'precio_total':
            solpeds = solpeds.order_by('total_price')
        elif order_by == 'fecha_resolucion':
            solpeds = solpeds.order_by('resolution_deadline')
        elif order_by == 'estado':
            solpeds = solpeds.order_by('status')
        else:
            solpeds = solpeds.order_by('createdAt')


    if filter_params['type'] and filter_params['input']:
        type_filter = filter_params['type']
        input_filter = filter_params['input']

        if type_filter == 'numero_solped':
            solpeds = solpeds.filter(_id__contains=input_filter) #works
        elif type_filter == 'codigo_material':
            solpeds = solpeds.filter(solpeditem__product__reference_code__exact=input_filter) #works
        elif type_filter == 'usuario_creador':
            solpeds = solpeds.filter(creator_user=input_filter) # filtrar por nombre
        elif type_filter == 'ubicacion':
            solpeds = solpeds.filter()
        elif type_filter == 'fecha_creacion':
            input_date = datetime.strptime(input_filter, '%Y-%m-%d')
            input_year = input_date.year
            input_month = input_date.month
            input_day = input_date.day
            solpeds = solpeds.filter(createdAt__year=input_year, createdAt__month=input_month, createdAt__day=input_day)
        elif type_filter == 'centro_costos':
            solpeds = solpeds.filter()
        elif type_filter == 'estado':
            solpeds = solpeds.filter(status=int(input_filter))

    # Aplicar paginación
    paginator = PageNumberPagination()
    paginator.page_size = 10  # Número de elementos por página
    result_page = paginator.paginate_queryset(solpeds, request)

    solpeds_serializer = SolpedSerializer(result_page, many=True)
    return paginator.get_paginated_response({'solpeds': solpeds_serializer.data, 'pages': paginator.page.paginator.num_pages})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_orders_by_user(request):
    solpeds = Solped.objects.filter(creator_user=request.user.pk)
    if solpeds is None:
        return Response(status=status.HTTP_404_NOT_FOUND)
    solpeds_serializer = SolpedSerializer(solpeds, many=True)
    return Response(solpeds_serializer.data)


#Se pide el id del producto de la solped
@api_view(['GET'])
@permission_classes([])
def get_documents_for_item(request, pk):
    solped_item_documents = Document.objects.filter(solped_item=pk)
    document_serializer = DocumentSerializer(solped_item_documents, many=True)
    return Response(document_serializer.data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([])
def get_documents_for_solped(request, pk):
    try:
        solped_item_documents = Document.objects.filter(solped=pk)
        document_serializer = DocumentSerializer(solped_item_documents, many=True)

        data = {'documents': document_serializer.data}
        return Response(data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': 'Error retrieving documents'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([])
def upload_document_item(request, pk):
    # Obtener el archivo del formulario enviado en la solicitud
    file = request.FILES.get('file', None)
    if not file:
        return Response({'error': 'No se proporcionó ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)

    # Asociar el documento con el solped_item especificado en la solicitud
    solped_item = SolpedItem.objects.get(pk=pk)
    solped = Solped.objects.get(pk=solped_item.solped.pk)
    Document.objects.create(solped_item=solped_item, document=file, solped=solped)

    return Response({'message': 'Archivo guardado exitosamente.'}, status=status.HTTP_201_CREATED)


